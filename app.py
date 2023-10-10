from flask import Flask, render_template, request, send_file
import openpyxl
import io
import re
import pdfplumber

app = Flask(__name__)

def extrair_informacoes(texto):
    informacoes = {
        "Codigo": [],
        "Descricao": [],
        "Class. Fiscal": [],
        "Preco Unit": [],
        "Valor IPI": []
    }

    for chave, valor in informacoes.items():
        matches = re.findall(rf'{chave}: (.+)', texto)
        informacoes[chave].extend(matches)

    return informacoes

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    manufacturer = request.form['manufacturer']
    pdf_file = request.files['pdf_file']

    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1).value = "Manufacturer"
    worksheet.cell(row=2, column=1).value = manufacturer

    worksheet.cell(row=1, column=2).value = "Mfr Part # (*)"
    worksheet.cell(row=2, column=2).value = ""
    
    worksheet.cell(row=1, column=3).value = "Description(*)"
    worksheet.cell(row=2, column=3).value = ""

    worksheet.cell(row=1, column=4).value = "Item Type (1=Product; 2=Managed Service; 3=Professional Service; 4=OEM Professional Service; 5=Software Licenses; 6=Third Party Professional Service; 7=OEM Maintenance; 8=Software; 9=Third Party Maintenance Services; 10=Logistics Charge; 11=Customs And Duties; 12=Training; 100=Pre-Payment; 999=New Section)"
    worksheet.cell(row=2, column=4).value = ""

    worksheet.cell(row=1, column=5).value = "Quantity(*)"
    worksheet.cell(row=2, column=5).value = ""

    worksheet.cell(row=1, column=6).value = "List Price"
    worksheet.cell(row=2, column=6).value = ""

    worksheet.cell(row=1, column=7).value = "Unit Price(*)"
    worksheet.cell(row=2, column=7).value = ""

    worksheet.cell(row=1, column=8).value = "Our Cost(*)"
    worksheet.cell(row=2, column=8).value = ""

    worksheet.cell(row=1, column=9).value = "UNSPSC"
    worksheet.cell(row=2, column=9).value = ""

    worksheet.cell(row=1, column=10).value = "External Comments"
    worksheet.cell(row=2, column=10).value = ""

    worksheet.cell(row=1, column=11).value = "Internal Comments"
    worksheet.cell(row=2, column=11).value = ""

    worksheet.cell(row=1, column=12).value = "Price Rule('Fixed' , 'Margin', 'Discount')"
    worksheet.cell(row=2, column=12).value = "Fixed"

    worksheet.cell(row=1, column=13).value = "Cost Factor 1"
    worksheet.cell(row=2, column=13).value = "0"

    worksheet.cell(row=1, column=14).value = "Cost Factor 2"
    worksheet.cell(row=2, column=14).value = "0"

    worksheet.cell(row=1, column=15).value = "Surcharges"
    worksheet.cell(row=2, column=15).value = ""

    worksheet.cell(row=1, column=16).value = "Vendor Maintenance"
    worksheet.cell(row=2, column=16).value = ""

    worksheet.cell(row=1, column=17).value = "Local Maintenance"
    worksheet.cell(row=2, column=17).value = ""

    worksheet.cell(row=1, column=18).value = "Currency('AED'=Arab Emirates Dirham, 'AUD'=Australian Dollar, 'BHD'=Bahraini Dinar, 'BWP'=Botswanan Pula, 'BRL'=Brazilean Real, 'GBP'=British Pound, 'CAD'=Canadian Dollar, 'CNY'=Chinese Renminbi, 'CZK'=Czech Krona, 'DKK'=Danish Krone, 'EUR'=Euro, 'HKD'=Hong Kong Dollar, 'INR'=Indian Rupee, 'IDR'=Indonesian Rupiah, 'JPY'=Japanese Yen, 'KES'=Kenyan Shilling, 'MYR'=Malaysian Ringgit, 'TWD'=New Taiwan Dollar, 'NZD'=New Zealand Dollar, 'NGN'=Nigerian Naira, 'NOK'=Norwegian Krone, 'PHP'=Philippine Peso, 'QAR'=Qatari Rial, 'SAR'=Saudi Riyals, 'SGD'=Singapore Dollar, 'ZAR'=South African Rand, 'KRW'=South Korean Won, 'SEK'=Swedish Krona, 'CHF'=Swiss Franc, 'THB'=Thailand Baht, 'USD'=US Dollar, 'VND'=Vietnamese Dong,)"
    worksheet.cell(row=2, column=18).value = "BRL"

    worksheet.cell(row=1, column=19).value = "Required Section"
    worksheet.cell(row=2, column=19).value = ""

    worksheet.cell(row=1, column=20).value = "Solution Type"
    worksheet.cell(row=2, column=20).value = ""

    worksheet.cell(row=1, column=21).value = "Preferred Supplier"
    worksheet.cell(row=2, column=21).value = "Others"

    worksheet.cell(row=1, column=22).value = "UOM"
    worksheet.cell(row=2, column=22).value = ""

    worksheet.cell(row=1, column=23).value = "Brazil NCM Code"
    worksheet.cell(row=2, column=23).value = "0"

    worksheet.cell(row=1, column=24).value = "Cost Factor 3"
    worksheet.cell(row=2, column=24).value = "0"

    worksheet.cell(row=1, column=25).value = "Cost Factor 4"
    worksheet.cell(row=2, column=25).value = "0"

    worksheet.cell(row=1, column=26).value = "Cost Factor 5"
    worksheet.cell(row=2, column=26).value = "0"
    
    worksheet.cell(row=1, column=27).value = "DealType"
    worksheet.cell(row=2, column=27).value = ""

    worksheet.cell(row=1, column=28).value = "DealValue"
    worksheet.cell(row=2, column=28).value = ""

    worksheet.cell(row=1, column=29).value = "SubItemType"
    worksheet.cell(row=2, column=29).value = ""

    worksheet.cell(row=1, column=30).value = "CategoryCode"
    worksheet.cell(row=2, column=30).value = ""

    worksheet.cell(row=1, column=31).value = "VendorQuoteNumber"
    worksheet.cell(row=2, column=31).value = ""

    with pdfplumber.open(pdf_file) as pdf:
        texto_completo = ''
        for pagina in pdf.pages:
            texto_completo += pagina.extract_text()

    informacoes = extrair_informacoes(texto_completo)

    for chave, valores in informacoes.items():
        for i, valor in enumerate(valores, start=2):
            worksheet[f'A{i}'] = chave
            worksheet[f'B{i}'] = valor

    workbook.save(output)

    output.seek(0)   

    # Salvando o arquivo temporário
    output_path = 'temp_results.xlsx'
    with open(output_path, 'wb') as temp_file:
        temp_file.write(output.read())

    return send_file(output_path, as_attachment=True, download_name='results.xlsx')

if __name__ == '__main__':
    app.run(debug=True)
