from flask import Flask, render_template, request, send_file
import openpyxl
import io
import re
import pdfplumber

app = Flask(__name__)

def extrair_informacoes(texto):
    informacoes = {
        "Codigo": [],
        "Descricao": [],
        "Class. Fiscal": [],
        "Preco Unit": [],
        "Valor IPI": []
    }

    for chave, valor in informacoes.items():
        matches = re.findall(rf'{chave}: (.+)', texto)
        informacoes[chave].extend(matches)

    return informacoes

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    manufacturer = request.form['manufacturer']
    pdf_file = request.files['pdf_file']

    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1).value = "Manufacturer"
    worksheet.cell(row=2, column=1).value = manufacturer

    worksheet.cell(row=1, column=2).value = "Mfr Part # (*)"
    worksheet.cell(row=2, column=2).value = manufacturer


    with pdfplumber.open(pdf_file) as pdf:
        first_page = pdf.pages[0]
        teste = first_page.extract_text()
        teste_data = re.search("([0-9]{2}/[0-9]{2}/[0-9]{4})", teste)

    tables = first_page.extract_tables(table_settings={"vertical_strategy": "lines", "horizontal_strategy": "lines"})

    operacoes = []
    #Itera as tabelas encontradas
    for table in tables:
      print(table[0])
      #print('segundo', table[0][0])

      #Aqui tem o pulo do gato pra identificar se a tabela é a tabela das operações kkk
      #Meu mestre de python chega tremer
      if table[0][0]== '':
        print(table[0])
        #Estrutura os dados encotrados na tabela
        operacao = {
          'tipo_operacao' : table[0][2],
          'mercado' : table[0][3],
          'nome_ativo' : table[0][5],
          'quantidade' : int(table[0][7]),
          'pm' : table[0][8],
          'valor_total' : table[0][9],
          'data' : data_pregao.group(1)
        }
        operacoes.append(operacao)
    return operacoes


    workbook.save(output)

    output.seek(0)   

    # Salvando o arquivo temporário
    output_path = 'temp_results.xlsx'
    with open(output_path, 'wb') as temp_file:
        temp_file.write(output.read())

    return send_file(output_path, as_attachment=True, download_name='results.xlsx')

if __name__ == '__main__':
    app.run(debug=True)
